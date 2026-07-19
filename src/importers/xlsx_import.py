"""Import XLSX klientów po external_id z podglądem zmian (WORKFLOW.md).

Ponowny import aktualizuje wyłącznie dane podstawowe; zadania, kontakty,
szkolenia, notatki i zdjęcie pozostają bez zmian. Brak klienta w pliku nie
usuwa go z bazy.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from models.entities import Client

# Nagłówek w pliku (po normalizacji) -> atrybut encji Client.
HEADER_MAP: dict[str, str] = {
    "id klienta": "external_id",
    "asii lp.": "external_id",
    "asii lp": "external_id",
    "external_id": "external_id",
    "id": "external_id",
    "imie": "first_name",
    "imię": "first_name",
    "nazwisko": "last_name",
    "telefon": "phone",
    "e-mail": "email",
    "email": "email",
    "data rekrutacji": "recruitment_date",
    "data ipd": "ipd_date",
    "cv": "cv_status",
    "zatrudnienie": "employment_status",
    "staz": "internship_status",
    "staż": "internship_status",
    "dz": "dz",
    "jc": "jc",
    "rp": "rp",
    "psycholog": "psychologist",
    "prawnik": "lawyer",
    "plec": "gender",
    "płeć": "gender",
    "stopien niepelnosprawnosci": "disability_degree",
    "stopień niepełnosprawności": "disability_degree",
    "symbol": "disability_symbol",
    "symbole sprzezone": "combined_symbols",
    "symbole sprzężone": "combined_symbols",
    "wyksztalcenie": "education",
    "wykształcenie": "education",
    "data waznosci orzeczenia": "certificate_valid_until",
    "data ważności orzeczenia": "certificate_valid_until",
    "poszukiwana praca": "desired_job",
    "komentarz": "import_comment",
    # długi nagłówek-instrukcja z realnych plików (symbole sprzężone)
    "jeśli niepełnosprawność sprzężona wpisać symbole ręcznie": "combined_symbols",
    "jesli niepelnosprawnosc sprzezona wpisac symbole recznie": "combined_symbols",
}

# dopasowanie po fragmencie nagłówka (gdy tekst nie jest dokładnie taki jak wyżej)
HEADER_CONTAINS = [
    ("sprzężon", "combined_symbols"),
    ("sprzezon", "combined_symbols"),
    ("stopień niepe", "disability_degree"),
    ("stopien niepe", "disability_degree"),
    ("ważności orzecz", "certificate_valid_until"),
    ("waznosci orzecz", "certificate_valid_until"),
    ("data rekrutacji", "recruitment_date"),
    ("data ipd", "ipd_date"),
    ("poszukiwana praca", "desired_job"),
    ("wykształ", "education"),
    ("wyksztal", "education"),
]

# Rozpoznawanie kolumny ID: „mocne" nagłówki (jednoznaczne) vs „słabe" (numer porządkowy).
# Gdy w pliku jest mocny nagłówek — używamy go; w innym wypadku bierzemy słaby (LP/Nr/…).
_ID_STRONG_EXACT = {"id klienta", "external_id", "asii lp.", "asii lp", "id", "identyfikator"}
_ID_STRONG_CONTAINS = ("asii", "id klient", "identyfikator")
_ID_WEAK_EXACT = {
    "lp", "lp.", "l.p.", "l. p.", "nr", "nr.", "numer", "poz", "poz.", "pozycja", "l/p",
}
_ID_RECOGNIZED = "ASII LP., ID klienta, LP, Nr, Numer, Poz., Identyfikator"


def _looks_like_lp(key: str) -> bool:
    """Nagłówek typu „<kod projektu> LP.” (np. ASII LP., AZ LP.) lub samo „LP” — to numer/ID."""
    tokens = key.replace(".", " ").split()
    return bool(tokens) and tokens[-1] == "lp"


def _is_strong_id(key: str) -> bool:
    return bool(key) and (key in _ID_STRONG_EXACT or any(f in key for f in _ID_STRONG_CONTAINS))


def _is_weak_id(key: str) -> bool:
    # numer porządkowy: „LP”, „<prefiks> LP.” (ASII/AZ/…), „Nr”, „Numer”, „Poz.”
    return key in _ID_WEAK_EXACT or _looks_like_lp(key)


DATE_FIELDS = {"recruitment_date", "ipd_date", "certificate_valid_until"}
# pola aktualizowane przy ponownym imporcie (dane podstawowe)
UPDATABLE_FIELDS = [v for v in dict.fromkeys(HEADER_MAP.values()) if v != "external_id"]


@dataclass
class RowError:
    row_number: int
    external_id: str
    message: str


@dataclass
class ImportPreview:
    file_name: str
    new: list[Client] = field(default_factory=list)
    updated: list[tuple[Client, Client]] = field(default_factory=list)  # (istniejący, docelowy)
    unchanged: list[Client] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    duplicates: list[RowError] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.new) + len(self.updated) + len(self.unchanged) + len(self.errors) + len(self.duplicates)


def _clean_id(value) -> str:
    """Czyści ID: liczby (także float „1.0” z Excela) do całkowitego napisu; przycina spacje."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _norm(value) -> str:
    """Normalizuje nagłówek: zwija białe znaki (w tym łamanie wiersza) do pojedynczej spacji."""
    if value is None:
        return ""
    return " ".join(str(value).split()).strip().lower()


def _match_header(name) -> str | None:
    """Zwraca atrybut encji dla nagłówka: dokładnie lub po fragmencie tekstu."""
    key = _norm(name)
    if not key:
        return None
    if key in HEADER_MAP:
        return HEADER_MAP[key]
    for fragment, field_name in HEADER_CONTAINS:
        if fragment in key:
            return field_name
    return None


import re

# data na początku komórki: d.m.rrrr / dd-mm-rrrr / rrrr-mm-dd (reszta, np. zakres godzin, ignorowana)
_DATE_DMY = re.compile(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})")
_DATE_YMD = re.compile(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})")


def _parse_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # wyłuskaj datę z początku komórki (dopuszcza dołączony zakres godzin, np. „01.04.2026 7:00-9:00”)
    m = _DATE_YMD.match(text) or _DATE_DMY.match(text)
    if not m:
        m = _DATE_YMD.search(text) or _DATE_DMY.search(text)
    if m:
        try:
            if m.re is _DATE_YMD:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            else:
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, d)
        except ValueError:
            pass
    raise ValueError(f"nieprawidłowa data: {text!r}")


def _norm_status(field: str, value) -> Optional[str]:
    text = _norm(value)
    if text == "":
        return None
    if field == "cv_status":
        # trójstan: brak / do poprawy / aktualne (rozpoznaje realne wpisy z bazy)
        if "popraw" in text:
            return "do_poprawy"
        if "brak" in text or text in ("nie", "0", "-"):
            return "brak"
        if "nieaktual" in text:  # wartość zgodności wstecz
            return "do_poprawy"
        if "aktual" in text:
            return "aktualne"
        return "do_poprawy"
    if field == "employment_status":
        return "zatrudniony" if text in ("zatrudniony", "zatrudniona", "tak", "1") else "bez_pracy"
    if field == "internship_status":
        return "w_trakcie" if text in ("w trakcie", "w_trakcie", "trwa") else "brak"
    return None


def parse_workbook(path: str | Path) -> tuple[list[dict], list[RowError]]:
    """Zwraca (wiersze jako słowniki pól, błędy). Wiersz zawiera _row i _values."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], [RowError(1, "", "Pusty plik.")]

    col_field: dict[int, str] = {}
    id_strong_idx: int | None = None
    id_weak_idx: int | None = None
    for idx, name in enumerate(header):
        key = _norm(name)
        # ID rozpoznajemy osobno: preferuj „mocny” nagłówek, w razie braku weź „słaby” (LP/Nr/…)
        if _is_strong_id(key):
            if id_strong_idx is None:
                id_strong_idx = idx
            continue
        if _is_weak_id(key):
            if id_weak_idx is None:
                id_weak_idx = idx
            continue
        field_name = _match_header(name)
        if field_name is not None and field_name != "external_id" and field_name not in col_field.values():
            col_field[idx] = field_name

    id_idx = id_strong_idx if id_strong_idx is not None else id_weak_idx
    if id_idx is not None:
        col_field[id_idx] = "external_id"
    if "external_id" not in col_field.values():
        found = [str(h).replace("\n", " ").strip() for h in header if h not in (None, "")]
        found_txt = ", ".join(f"„{h}”" for h in found) if found else "(brak nagłówków)"
        return [], [RowError(
            1, "",
            "Brak kolumny z ID klienta.\n"
            f"Znalezione nagłówki: {found_txt}.\n"
            f"Rozpoznawane nazwy ID: {_ID_RECOGNIZED}.",
        )]

    parsed: list[dict] = []
    errors: list[RowError] = []
    for r_idx, raw in enumerate(rows_iter, start=2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        values: dict[str, object] = {}
        for idx, field_name in col_field.items():
            cell = raw[idx] if idx < len(raw) else None
            is_empty = cell is None or str(cell).strip() == ""
            if field_name in DATE_FIELDS:
                if is_empty:
                    continue  # pusta komórka nie nadpisuje istniejącej daty
                try:
                    values[field_name] = _parse_date(cell)
                except ValueError:
                    # błędna data NIE może wyrzucać całego klienta — pomijamy tylko to pole
                    continue
            elif field_name in ("cv_status", "employment_status", "internship_status"):
                norm = _norm_status(field_name, cell)
                if norm is not None:
                    values[field_name] = norm
            else:
                if not is_empty:  # pusta komórka nie kasuje istniejącej wartości
                    values[field_name] = str(cell).strip()
        ext = _clean_id(values.get("external_id", ""))
        if not ext:
            errors.append(RowError(r_idx, "", "Brak ID klienta w wierszu."))
            continue
        values["external_id"] = ext
        values["_row"] = r_idx
        parsed.append(values)

    wb.close()
    return parsed, errors


def build_preview(store, path: str | Path) -> ImportPreview:
    path = Path(path)
    parsed, errors = parse_workbook(path)
    preview = ImportPreview(file_name=path.name, errors=errors)

    seen: dict[str, int] = {}
    for values in parsed:
        ext = values["external_id"]
        if ext in seen:
            preview.duplicates.append(
                RowError(values["_row"], ext, f"Duplikat ID w pliku (także wiersz {seen[ext]}).")
            )
            continue
        seen[ext] = values["_row"]

        existing = store.find_by_external_id(ext)
        if existing is None:
            has_name = bool(
                str(values.get("first_name", "")).strip()
                or str(values.get("last_name", "")).strip()
            )
            has_data = any(k not in ("external_id", "_row") for k in values)
            # wiersz z samym numerem LP (bez nazwiska i bez danych) to „widmo" z arkusza —
            # pomijamy, żeby nie tworzyć pustych rekordów zawyżających licznik klientów
            if not has_name and not has_data:
                continue
            # rekord z danymi, ale bez nazwiska — zostaje, z czytelnym placeholderem
            if not has_name:
                values["last_name"] = "(bez nazwiska)"
            preview.new.append(_client_from_values(values))
        else:
            target = _apply_values(existing, values)
            if _basic_differs(existing, target):
                preview.updated.append((existing, target))
            else:
                preview.unchanged.append(existing)
    return preview


def _client_from_values(values: dict) -> Client:
    client = Client(
        id=0,
        external_id=values["external_id"],
        first_name=str(values.get("first_name", "")).strip(),
        last_name=str(values.get("last_name", "")).strip(),
    )
    return _apply_values(client, values)


def _apply_values(base: Client, values: dict) -> Client:
    from copy import copy

    target = copy(base)
    for f in UPDATABLE_FIELDS:
        if f in values:
            setattr(target, f, values[f])
    return target


def _basic_differs(a: Client, b: Client) -> bool:
    return any(getattr(a, f) != getattr(b, f) for f in UPDATABLE_FIELDS)
