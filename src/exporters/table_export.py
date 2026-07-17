"""Eksport prostej tabeli (nagłówki + wiersze) do CSV / XLSX / PDF.

CSV — biblioteka standardowa; XLSX — openpyxl; PDF — reportlab (BUILD.md).
Pliki trafiają do data/exports/.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from config import data_dir


def exports_dir() -> Path:
    path = data_dir() / "exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


@dataclass
class TableData:
    title: str
    headers: list[str]
    rows: list[list[str]]


def _timestamped(prefix: str, ext: str) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return exports_dir() / f"{prefix}_{stamp}.{ext}"


def export_csv(table: TableData, prefix: str = "analityka") -> Path:
    path = _timestamped(prefix, "csv")
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(table.headers)
        writer.writerows(table.rows)
    return path


def export_xlsx(table: TableData, prefix: str = "analityka") -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = _timestamped(prefix, "xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = (table.title or "Analityka")[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D2330")
    ws.append(table.headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for row in table.rows:
        ws.append(row)

    for col_idx, header in enumerate(table.headers, start=1):
        width = len(str(header))
        for row in table.rows:
            if col_idx - 1 < len(row):
                width = max(width, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(width + 4, 60)
    ws.freeze_panes = "A2"

    wb.save(str(path))
    return path


def export_pdf(table: TableData, prefix: str = "analityka") -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from exporters.fonts import FONT_BOLD, register_pdf_fonts

    font = register_pdf_fonts()
    bold = FONT_BOLD if font != "Helvetica" else "Helvetica-Bold"
    path = _timestamped(prefix, "pdf")
    doc = SimpleDocTemplate(
        str(path), pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    for name in ("Title", "Normal", "BodyText", "Heading3"):
        styles[name].fontName = font
    styles["Title"].fontName = bold
    story = [
        Paragraph(table.title or "Analityka", styles["Title"]),
        Paragraph(
            "Wygenerowano: " + datetime.now().strftime("%d.%m.%Y %H:%M"), styles["Normal"]
        ),
        Spacer(1, 0.5 * cm),
    ]

    cell_style = styles["BodyText"]
    cell_style.fontSize = 8
    cell_style.leading = 10
    data = [[Paragraph(f"<b>{h}</b>", cell_style) for h in table.headers]]
    for row in table.rows:
        data.append([Paragraph(str(c), cell_style) for c in row])

    pdf_table = Table(data, repeatRows=1)
    pdf_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D2330")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B6C0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F2F6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(pdf_table)
    doc.build(story)
    return path
