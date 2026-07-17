"""Generuje szablon importu XLSX z realnymi nagłówkami oraz plik testowy 150 klientów.

Użycie:
    python tools/make_import_template.py           # szablon (pusty) + plik 150 wierszy
Pliki trafiają do docs/ (szablon) i do katalogu bieżącego (plik testowy).
"""
from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent

# Nagłówki zgodne z realnym plikiem użytkownika (kolejność jak w bazie).
HEADERS = [
    "ASII\nLP.", "IMIĘ", "NAZWISKO", "DATA REKRUTACJI", "DATA IPD", "CV", "ZATRUDNIENIE",
    "DZ", "JC", "RP", "PŁEĆ", "STOPIEŃ NIEPEŁNOSPRAWNOŚCI", "SYMBOL",
    "jeśli niepełnosprawność sprzężona wpisać symbole ręcznie", "WYKSZTAŁCENIE",
    "DATA WAŻNOŚCI ORZECZENIA", "POSZUKIWANA PRACA", "KOMENTARZ",
]

IMIONA_K = ["Anna", "Katarzyna", "Maria", "Ewa", "Zofia", "Magdalena", "Agnieszka", "Barbara"]
IMIONA_M = ["Jan", "Piotr", "Andrzej", "Tomasz", "Marek", "Paweł", "Krzysztof", "Michał"]
NAZWISKA = ["Kowalski", "Nowak", "Wiśniewski", "Wójcik", "Kowalczyk", "Kamiński", "Lewandowski",
            "Zieliński", "Szymański", "Woźniak", "Dąbrowski", "Kozłowski", "Mazur", "Kwiatkowski"]
STOPNIE = ["Lekki", "Umiarkowany", "Znaczny"]
SYMBOLE = ["01-U", "02-P", "03-L", "04-O", "05-R", "07-S", "10-N", "11-I"]
WYKSZTALCENIE = ["Podstawowe", "Zawodowe", "Średnie", "Wyższe"]
PRACE = ["Pracownik biurowy", "Magazynier", "Ochrona mienia", "Sprzedawca", "Księgowość",
         "Grafik komputerowy", "Kierowca kat. B", "Recepcja"]


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1D6F42")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40


def build_template() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Klienci"
    ws.append(HEADERS)
    _style_header(ws)
    # jeden przykładowy wiersz jako podpowiedź formatu
    ws.append([
        "AS-0001", "Jan", "Przykładowy", "2026-01-15", "2026-02-01", "aktualne", "bez pracy",
        "Tak", "Nie", "Nie", "Mężczyzna", "Umiarkowany", "05-R", "07-S", "Średnie",
        "2027-12-31", "Magazynier", "Wiersz przykładowy — usuń przed importem",
    ])
    out = ROOT / "docs" / "szablon_importu.xlsx"
    wb.save(out)
    return out


def build_test_150(path: Path | None = None) -> Path:
    rng = random.Random(42)
    wb = Workbook()
    ws = wb.active
    ws.title = "Klienci"
    ws.append(HEADERS)
    _style_header(ws)
    today = date.today()
    for i in range(1, 151):
        female = rng.random() < 0.5
        first = rng.choice(IMIONA_K if female else IMIONA_M)
        last = rng.choice(NAZWISKA) + ("a" if female and rng.random() < 0.5 else "")
        rec = today - timedelta(days=rng.randint(30, 400))
        ipd = rec + timedelta(days=rng.randint(5, 30))
        ws.append([
            f"AS-{3000 + i}", first, last, rec.isoformat(), ipd.isoformat(),
            rng.choice(["aktualne", "nieaktualne"]),
            rng.choice(["bez pracy", "zatrudniony"]),
            rng.choice(["Tak", "Nie"]), rng.choice(["Tak", "Nie"]), rng.choice(["Tak", "Nie"]),
            "Kobieta" if female else "Mężczyzna", rng.choice(STOPNIE), rng.choice(SYMBOLE),
            rng.choice(["", "07-S", "10-N"]), rng.choice(WYKSZTALCENIE),
            (today + timedelta(days=rng.randint(60, 900))).isoformat(),
            rng.choice(PRACE), "",
        ])
    out = path or (ROOT / "test_import_150.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    t = build_template()
    print(f"Szablon: {t}")
    f = build_test_150()
    print(f"Plik testowy (150 klientów): {f}")
